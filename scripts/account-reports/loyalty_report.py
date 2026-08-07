#!/usr/bin/env python3
"""
Xtresse account/loyalty report generator.

Turns a full Shopify orders CSV export into a branded multi-tab workbook:
  • Tier Progress  – every ship-to location that bought the target product in the
    period, by rep, with cases ordered, current/next tier, cases-to-next,
    phone/email, address, last order $, and cross-sell flags.
  • Lapsed Locations – ordered before the period but nothing in it (win-back list).
  • Validation     – 1 account per tier per rep with the actual orders behind the number.
  • Read Me        – tier rules, field guide, methodology.

WHY A CSV: the Shopify MCP/Admin connector caps the Orders API at ~60 days
(no read_all_orders). The Admin UI export is NOT capped, has full history, every
line item, order tags (= rep), and shipping + billing. So the workflow is:

  1. In Shopify Admin → Orders → Export → "All orders" → "Plain CSV file" → Export.
     (Drop it in ~/Downloads.)
  2. python3 loyalty_report.py "<orders_export.csv>" "<output.xlsx>"

Reusable for ANY product: change CONFIG.target (SKU + label), period, and tiers.
e.g. Serum Case (X-FRC-30ML-CASE), XVIE (X-XVIE-2ML-006), a different quarter, etc.

Identity note: a "location" is Shipping Company + ZIP (the real ship-to clinic).
Do NOT key on billing company (group/AP entities span clinics) or ZIP alone
(one ZIP can hold several clinics). ADCS is excluded (canonical), cancelled
orders excluded, rep comes from the order tags.
"""
import csv, re, json, sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================== CONFIG ==============================
CONFIG = {
    # The loyalty product: any line item with one of these SKUs counts as 1 "case".
    "target_skus": {"860011740100"},
    "target_label": "Gummy Case",
    "unit_word": "case",                      # shown in headers / read-me
    "period": ("2026-04-01", "2026-06-30"),   # inclusive shop-local dates
    "period_label": "Q2 2026 (Apr 1 – Jun 30)",
    # Tier thresholds (min units, name), ascending. Next-tier / cases-to-next derived.
    "tiers": [(1, "Silver"), (2, "Gold"), (4, "Platinum"), (6, "Diamond")],
    # Cross-sell "has this location ever ordered X?" flags. Match by sku OR title keyword.
    "crosssell": {
        "Serum": {"skus": {"X-FRC-30ML-CASE"}, "keywords": ["serum"]},
        "XVIE":  {"skus": set(), "keywords": ["xvie"]},
    },
    # Order tag that grants a bonus item (e.g. an extra serum) regardless of tier.
    "bonus_tag": "JUNETIERUP",
    "bonus_label": "Bonus Serum (JUNETIERUP)",
    # Tiers whose accounts get a VIP Kit + Pull-Up Banner (per the Q2 thank-you card art).
    "kit_tiers": {"Platinum", "Diamond"},
    "title": "Xtresse — B2B Loyalty Report",
}
ROSTER = ["Jamie Bergeron","Michelle Spencer","Dia Lamport","Cheryl Greiber","Denisse Schimelpfening",
    "Tyler De Masi","Laura Mann","Sherry Quinn","Michelle Boehle","Sonia Mace","Taylor Bates","Julie Fetter",
    "Becky Curry","Ryan Masa","Heidi Fisher","Gina Napoli","Amy Pierre","Megan Gilbert","Bridget Selberg",
    "Carrie Dodge","Morgan Hood","James Tuckett","Lexi Cavaliere","Jim & Anne Weeks","Sevi McCutcheon","Krista Taylor"]
# Brand palette / fonts
MAROON="5C2F2E"; CREAM="F8F0E2"; INK="2D1F1A"; SAGE="3E6B50"; FONT="Arial"
# ====================================================================

PROV={'Alabama':'AL','Alaska':'AK','Arizona':'AZ','Arkansas':'AR','California':'CA','Colorado':'CO','Connecticut':'CT','Delaware':'DE','Florida':'FL','Georgia':'GA','Hawaii':'HI','Idaho':'ID','Illinois':'IL','Indiana':'IN','Iowa':'IA','Kansas':'KS','Kentucky':'KY','Louisiana':'LA','Maine':'ME','Maryland':'MD','Massachusetts':'MA','Michigan':'MI','Minnesota':'MN','Mississippi':'MS','Missouri':'MO','Montana':'MT','Nebraska':'NE','Nevada':'NV','New Hampshire':'NH','New Jersey':'NJ','New Mexico':'NM','New York':'NY','North Carolina':'NC','North Dakota':'ND','Ohio':'OH','Oklahoma':'OK','Oregon':'OR','Pennsylvania':'PA','Rhode Island':'RI','South Carolina':'SC','South Dakota':'SD','Tennessee':'TN','Texas':'TX','Utah':'UT','Vermont':'VT','Virginia':'VA','Washington':'WA','West Virginia':'WV','Wisconsin':'WI','Wyoming':'WY','District of Columbia':'DC'}

def norm(s): return re.sub(r'\s+',' ',re.sub(r'[^a-z0-9 ]+',' ',str(s or '').lower().replace('&',' and '))).strip()
def money(s):
    try: return float(re.sub(r'[^0-9.\-]','',s or "0") or 0)
    except: return 0.0
def fmtphone(p):
    d=re.sub(r'\D','',(p or "").lstrip("'"))
    if len(d)==11 and d[0]=='1': d=d[1:]
    return f"({d[0:3]}) {d[3:6]}-{d[6:]}" if len(d)==10 else (p or "").lstrip("'").strip()

ROSTER_N={norm(r):r for r in ROSTER}
def find_rep(tags):
    for t in tags.split(','):
        if norm(t) in ROSTER_N: return ROSTER_N[norm(t)]
    return None
def is_adcs(tags): return ('advanced derm' in tags.lower()) or ('adcs' in [norm(t) for t in tags.split(',')])
def has_tag(tags, tag): return norm(tag) in [norm(t) for t in tags.split(',')]

def tier_of(c, tiers):
    name="-"
    for thr,nm in tiers:
        if c>=thr: name=nm
    return name
def next_of(c, tiers):
    for thr,nm in tiers:
        if c<thr: return nm, thr-c
    return "-", 0   # top tier

def load_orders(path):
    """Shopify CSV export is multi-row per order (order fields on first row, extra
    line items as continuation rows sharing the Name). Group by Name."""
    O={}
    with open(path, encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            nm=row.get("Name")
            if nm is None: continue
            o=O.get(nm)
            if not o:
                o=dict(date="",tags="",cancelled=False,shipCo="",shipName="",addr1="",addr2="",city="",
                       prov="",zip="",phone="",bphone="",email="",total="",units=0,cross={k:False for k in CONFIG["crosssell"]})
                O[nm]=o
            if row.get("Created at"): o["date"]=row["Created at"][:10]
            if row.get("Tags"): o["tags"]=row["Tags"]
            if (row.get("Cancelled at") or "").strip(): o["cancelled"]=True
            for k,c in [("shipCo","Shipping Company"),("shipName","Shipping Name"),("addr1","Shipping Address1"),
                        ("addr2","Shipping Address2"),("city","Shipping City"),("prov","Shipping Province Name"),
                        ("zip","Shipping Zip"),("phone","Shipping Phone"),("bphone","Billing Phone"),
                        ("email","Email"),("total","Total")]:
                v=(row.get(c) or "").strip()
                if v and not o[k]: o[k]=v
            sku=(row.get("Lineitem sku") or "").strip(); name_l=(row.get("Lineitem name") or "").lower()
            try: q=int(float(row.get("Lineitem quantity") or 0))
            except: q=0
            if sku in CONFIG["target_skus"]: o["units"]+=q
            for label,spec in CONFIG["crosssell"].items():
                if sku in spec["skus"] or any(kw in name_l for kw in spec["keywords"]): o["cross"][label]=True
    return O

def load_orders_json(path):
    """Read the normalized JSON produced by pull_orders.mjs (direct Shopify pull)."""
    rows=json.load(open(path, encoding="utf-8"))
    O={}
    for o in rows:
        units=0; cross={k:False for k in CONFIG["crosssell"]}
        for li in o.get("lineItems",[]):
            sku=(li.get("sku") or "").strip(); name_l=(li.get("name") or "").lower(); q=int(li.get("qty") or 0)
            if sku in CONFIG["target_skus"]: units+=q
            for label,spec in CONFIG["crosssell"].items():
                if sku in spec["skus"] or any(kw in name_l for kw in spec["keywords"]): cross[label]=True
        O[o["name"]]=dict(date=(o.get("date") or "")[:10],tags=o.get("tags") or "",cancelled=bool(o.get("cancelled")),
            shipCo=o.get("shipCo") or "",shipName=o.get("shipName") or "",addr1=o.get("addr1") or "",addr2=o.get("addr2") or "",
            city=o.get("city") or "",prov=o.get("prov") or "",zip=o.get("zip") or "",phone=o.get("phone") or "",
            bphone=o.get("bphone") or "",email=o.get("email") or "",total=str(o.get("total") or "0"),units=units,cross=cross)
    return O

def load(path):
    return load_orders_json(path) if path.lower().endswith(".json") else load_orders(path)

# How the territory engine decided the rep (from assign_reps.mjs `basis`).
BASIS_LABEL={"tag":"Tagged","tag-area":"Tagged (area)","tag-stale":"Tagged (stale)",
    "zip":"Territory · ZIP","prefix":"Territory · ZIP area","region":"Territory · region",
    "state":"Territory · state","proximity":"Territory · nearest","override-zip":"Territory · assigned",
    "override-prefix":"Territory · assigned","override-state":"Territory · assigned","unassigned":"Unassigned"}

def build(path_csv, path_out, territory_path=None):
    P0,P1=CONFIG["period"]; tiers=CONFIG["tiers"]
    in_p=lambda d: P0<=d<=P1
    O=load(path_csv)
    tmap={}
    if territory_path:
        try: tmap=json.load(open(territory_path,encoding="utf-8"))
        except Exception: tmap={}
    acc={}
    for nm,o in O.items():
        if o["cancelled"] or not o["date"] or is_adcs(o["tags"]): continue
        rep=find_rep(o["tags"]); tl=[norm(t) for t in o["tags"].split(',')]
        if not (rep or 'b2b' in tl or 'wholesale' in o["tags"].lower()): continue   # B2B only
        label=o["shipCo"] or o["shipName"] or o["email"]
        if not label: continue
        z=(o["zip"] or "")[:5]; key=norm(label)+"|"+z
        a=acc.get(key)
        if not a:
            a=dict(key=key,label=label,zip=z,rep=None,repDate="",city=o["city"],state=PROV.get(o["prov"],o["prov"]),
                   addr1=o["addr1"],addr2=o["addr2"],phone="",email="",last="",lastVal=0,first=o["date"],
                   units=0,orders=0,hist=[],cross={k:False for k in CONFIG["crosssell"]},bonus=False)
            acc[key]=a
        if rep and o["date"]>=a["repDate"]: a["rep"]=rep; a["repDate"]=o["date"]
        if o["date"]>a["last"]:
            a.update(last=o["date"],lastVal=money(o["total"]),label=label,addr1=o["addr1"],addr2=o["addr2"],
                     city=o["city"],state=PROV.get(o["prov"],o["prov"]),zip=z)
            if o["phone"] or o["bphone"]: a["phone"]=o["phone"] or o["bphone"]
            if o["email"]: a["email"]=o["email"]
        if o["date"]<a["first"]: a["first"]=o["date"]
        if not a["phone"] and (o["phone"] or o["bphone"]): a["phone"]=o["phone"] or o["bphone"]
        if not a["email"] and o["email"]: a["email"]=o["email"]
        for k,v in o["cross"].items():
            if v: a["cross"][k]=True
        if in_p(o["date"]):
            a["orders"]+=1
            if o["units"]>0: a["units"]+=o["units"]; a["hist"].append({"date":o["date"],"name":nm,"qty":o["units"]})
            if has_tag(o["tags"], CONFIG["bonus_tag"]): a["bonus"]=True

    def addr(a):
        return ", ".join(p for p in [a["addr1"],a["addr2"],a["city"],a["state"],a["zip"]] if p)
    YN=lambda b:"Yes" if b else "No"
    tab1=[]; tab2=[]
    for a in acc.values():
        terr=tmap.get(a["key"],{})
        rep=terr.get("rep") or a["rep"] or "(Unassigned)"
        tagRep=terr.get("tagRep") or a["rep"] or ""
        via=BASIS_LABEL.get(terr.get("basis"), ("Tagged" if a["rep"] else "Unassigned"))
        openedBy=tagRep if (tagRep and tagRep!=rep) else ""
        b=dict(rep=rep,via=via,openedBy=openedBy,location=a["label"],address=addr(a),city=a["city"],state=a["state"],zip=a["zip"],
               phone=fmtphone(a["phone"]),email=a["email"],lastVal=round(a["lastVal"]),
               cross={k:YN(v) for k,v in a["cross"].items()})
        if a["units"]>=1:
            nt,ctn=next_of(a["units"],tiers); tier=tier_of(a["units"],tiers)
            tab1.append({**b,"units":a["units"],"tier":tier,"nextTier":nt,"casesToNext":ctn,
                         "bonus":a["bonus"],"kit":tier in CONFIG["kit_tiers"],
                         "hist":sorted(a["hist"],key=lambda x:x["date"])})
        elif a["orders"]==0 and a["first"]<P0:
            y,m,dd=int(a["last"][:4]),int(a["last"][5:7]),int(a["last"][8:10])
            ds=(date(*map(int,P1.split("-")))-date(y,m,dd)).days
            tab2.append({**b,"lastOrder":a["last"],"daysSince":ds})
    tab1.sort(key=lambda t:((t["rep"]=="(Unassigned)"),t["rep"],t["casesToNext"],-t["units"],t["location"]))
    tab2.sort(key=lambda t:t["lastOrder"],reverse=True); tab2.sort(key=lambda t:t["rep"])
    # validation sample: 1 per tier per rep
    byrep={}
    for t in tab1:
        if t["rep"]=="(Unassigned)": continue
        byrep.setdefault(t["rep"],{}).setdefault(t["tier"],[]).append(t)
    sample=[]
    for rep in sorted(byrep):
        for _,nm in tiers:
            l=sorted(byrep[rep].get(nm,[]),key=lambda x:-x["units"])
            if l: sample.append(l[0])
    write_xlsx(path_out, tab1, tab2, sample)
    return tab1, tab2, sample

# ----------------------------- xlsx -----------------------------
def write_xlsx(out, tab1, tab2, sample):
    cs=list(CONFIG["crosssell"].keys()); uw=CONFIG["unit_word"].title()
    hf=Font(name=FONT,bold=True,color=CREAM,size=11); tf=Font(name=FONT,bold=True,color=MAROON,size=18)
    sfn=Font(name=FONT,color=INK,size=10); bfn=Font(name=FONT,bold=True,color=INK,size=10); base=Font(name=FONT,color=INK,size=10)
    hfill=PatternFill("solid",fgColor=MAROON); band=PatternFill("solid",fgColor=CREAM); amber=PatternFill("solid",fgColor="F3E3C7")
    thin=Side(style="thin",color="D8C9B0"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
    C=Alignment("center",vertical="center"); L=Alignment("left",vertical="center"); Rg=Alignment("right",vertical="center")
    wb=Workbook()
    def header(ws,row,heads,widths):
        for c,(h,w) in enumerate(zip(heads,widths),1):
            x=ws.cell(row=row,column=c,value=h); x.font=hf; x.fill=hfill; x.alignment=C; x.border=bd
            ws.column_dimensions[get_column_letter(c)].width=w
        ws.freeze_panes=ws.cell(row=row+1,column=1)
    def noflag(cc,v):
        if v=="No": cc.font=Font(name=FONT,bold=True,color=SAGE,size=10)
    tiers_str="   ".join(f"{nm} = {thr}{'+' if (thr,nm)==CONFIG['tiers'][-1] else ''}" for thr,nm in CONFIG["tiers"])

    rm=wb.active; rm.title="Read Me"; rm.sheet_view.showGridLines=False
    rm["B2"]=CONFIG["title"]; rm["B2"].font=tf
    rm["B3"]=f'{CONFIG["period_label"]} · {CONFIG["target_label"]} · by ship-to location & rep'; rm["B3"].font=Font(name=FONT,color=MAROON,size=11,italic=True)
    tiers_n={}
    for t in tab1: tiers_n[t["tier"]]=tiers_n.get(t["tier"],0)+1
    lines=["",("HOW TO USE","h"),
     "Tab 1 — Tier Progress: every ship-to location that ordered the product this period, by rep,",
     "    with phone/email, units, current tier, and units to the next tier. Filter Rep to your name.",
     "    Amber rows = one unit from the next tier.",
     "Tab 2 — Lapsed Locations: ordered before the period but nothing in it — win-back list.",
     "Tab 3 — Validation: one account per tier per rep with the actual orders behind the number.",
     "",("REP ASSIGNMENT","h"),
     "• Rep = the CURRENT territory owner (sales-rep-dashboard logic): a fresh order tag (≤120d)",
     "    wins; otherwise the account flows to whoever owns the territory now (declared territory →",
     "    ZIP area → state → nearest rep).",
     "• Assigned Via = how the rep was decided (Tagged / Territory · state / ZIP / assigned / …).",
     "• Opened By (tag) = the rep tagged on the orders, shown only when it differs from the current",
     "    owner (i.e. the account was reassigned by territory). Tag = who opened it / commission credit.",
     "",("TIERS","h"),"    "+tiers_str,
     "",("BONUS & KITS","h"),
     f'• Bonus Serum = account has a Q2 order tagged "{CONFIG["bonus_tag"]}" (extra serum, regardless of tier).',
     f'• VIP Kit + Banner = {"/".join(sorted(CONFIG["kit_tiers"]))} tier accounts (per the Q2 thank-you card art).',
     "",("NOTES","h"),
     "• Rows = ship-to LOCATIONS (Shipping Company + ZIP); distinct clinics at one ZIP stay separate.",
     "• Built from the full Shopify order history. Cancelled excluded, B2B only, ADCS excluded.",
     "• National accounts are NOT split out here — handled on a separate list once confirmed.",
    ]
    r=5
    for it in lines:
        if isinstance(it,tuple): rm[f"B{r}"]=it[0]; rm[f"B{r}"].font=Font(name=FONT,bold=True,color=MAROON,size=12)
        else: rm[f"B{r}"]=it; rm[f"B{r}"].font=sfn
        r+=1
    r+=1; rm[f"B{r}"]="SNAPSHOT"; rm[f"B{r}"].font=Font(name=FONT,bold=True,color=MAROON,size=12); r+=1
    snap=[("Locations with orders this period",len(tab1))]+[(f"  • {nm}",tiers_n.get(nm,0)) for _,nm in CONFIG["tiers"]]
    snap+=[("One unit from next tier",sum(1 for t in tab1 if t["casesToNext"]==1))]
    snap+=[("Bonus Serum (JUNETIERUP)",sum(1 for t in tab1 if t["bonus"]))]
    snap+=[("VIP Kit + Banner (Platinum/Diamond)",sum(1 for t in tab1 if t["kit"]))]
    snap+=[("Reassigned by territory (tag ≠ current owner)",sum(1 for t in tab1 if t["openedBy"]))]
    snap+=[(f"Buyers who've NEVER ordered {c}",sum(1 for t in tab1 if t["cross"][c]=="No")) for c in cs]
    snap+=[("Lapsed locations",len(tab2))]
    for label,v in snap:
        rm[f"B{r}"]=label; rm[f"B{r}"].font=sfn; rm[f"C{r}"]=v; rm[f"C{r}"].font=bfn; rm[f"C{r}"].alignment=Rg; r+=1
    rm.column_dimensions["A"].width=3; rm.column_dimensions["B"].width=86; rm.column_dimensions["C"].width=12

    # Tab 1
    s1=wb.create_sheet("Tier Progress"); s1.sheet_view.showGridLines=False
    heads=["Rep","Assigned Via","Opened By (tag)","Location","Address","City","State","ZIP","Phone","Email",f"{uw}s This Period","Current Tier","Next Tier",f"{uw}s to Next Tier","Last Order ($)","Bonus Serum (JUNETIERUP)?","VIP Kit + Banner?"]+[f"{c}?" for c in cs]
    widths=[18,17,16,28,38,14,6,7,15,28,15,12,11,15,13,20,15]+[8]*len(cs)
    ncol=len(heads)
    s1["A1"]=f'Tier Progress — {CONFIG["period_label"]} {CONFIG["target_label"]} by Ship-To Location & Rep'; s1["A1"].font=tf; s1.merge_cells(f"A1:{get_column_letter(ncol)}1")
    header(s1,3,heads,widths); rr=4
    centercols={2,7,8,11,12,13,14,16,17}|{17+1+i for i in range(len(cs))}
    for t in tab1:
        vals=[t["rep"],t["via"],t["openedBy"],t["location"],t["address"],t["city"],t["state"],t["zip"],t["phone"],t["email"],t["units"],t["tier"],t["nextTier"],("Top tier" if t["casesToNext"]==0 else t["casesToNext"]),t["lastVal"],("Yes" if t["bonus"] else "No"),("Yes" if t["kit"] else "No")]+[t["cross"][c] for c in cs]
        one=t["casesToNext"]==1
        for c,v in enumerate(vals,1):
            cc=s1.cell(row=rr,column=c,value=v); cc.font=base; cc.border=bd
            cc.alignment=C if c in centercols else (Rg if c==15 else L)
            cc.fill=amber if one else (band if rr%2==0 else PatternFill())
            if c==15 and isinstance(v,(int,float)): cc.number_format='$#,##0'
            if c==14 and one: cc.font=Font(name=FONT,bold=True,color=MAROON,size=10)
            if c in (16,17) and v=="Yes": cc.font=Font(name=FONT,bold=True,color=MAROON,size=10)
            if c>17: noflag(cc,v)
        rr+=1
    s1.auto_filter.ref=f"A3:{get_column_letter(ncol)}{rr-1}"

    # Kits & Banners — fulfillment list: Platinum/Diamond ship-to locations
    kit_rows=[t for t in tab1 if t["kit"]]
    kit_rows.sort(key=lambda t:(t["rep"],t["tier"]!="Diamond",t["location"]))
    sk=wb.create_sheet("Kits & Banners"); sk.sheet_view.showGridLines=False
    headsk=["Rep","Location","Address","City","State","ZIP","Phone","Email","Current Tier",f"{uw}s This Period","Bonus Serum (JUNETIERUP)?"]
    widthsk=[18,28,38,14,6,7,15,28,12,15,20]; nk=len(headsk)
    sk["A1"]="VIP Kits & Pull-Up Banners — ship to every Platinum + Diamond location"; sk["A1"].font=tf; sk.merge_cells(f"A1:{get_column_letter(nk)}1")
    header(sk,3,headsk,widthsk); rr=4
    cck={6,7,9,10,11}
    for t in kit_rows:
        vals=[t["rep"],t["location"],t["address"],t["city"],t["state"],t["zip"],t["phone"],t["email"],t["tier"],t["units"],("Yes" if t["bonus"] else "No")]
        for c,v in enumerate(vals,1):
            cc=sk.cell(row=rr,column=c,value=v); cc.font=base; cc.border=bd
            cc.alignment=C if c in cck else L
            if rr%2==0: cc.fill=band
            if c==9: cc.font=Font(name=FONT,bold=True,color=(MAROON if v=="Diamond" else INK),size=10)
            if c==11 and v=="Yes": cc.font=Font(name=FONT,bold=True,color=MAROON,size=10)
        rr+=1
    sk.auto_filter.ref=f"A3:{get_column_letter(nk)}{rr-1}"

    # Tab 2
    s2=wb.create_sheet("Lapsed Locations"); s2.sheet_view.showGridLines=False
    heads2=["Rep","Assigned Via","Opened By (tag)","Location","Address","City","State","ZIP","Phone","Email","Last Order","Days Since","Last Order ($)"]+[f"{c}?" for c in cs]
    widths2=[18,17,16,28,38,14,6,7,15,28,12,11,13]+[8]*len(cs); n2=len(heads2)
    s2["A1"]="Lapsed Locations — ordered before the period, nothing in it yet"; s2["A1"].font=tf; s2.merge_cells(f"A1:{get_column_letter(n2)}1")
    header(s2,3,heads2,widths2); rr=4
    cc2={2,7,8,11,12}|{13+1+i for i in range(len(cs))}
    for t in tab2:
        vals=[t["rep"],t["via"],t["openedBy"],t["location"],t["address"],t["city"],t["state"],t["zip"],t["phone"],t["email"],t["lastOrder"],t["daysSince"],t["lastVal"]]+[t["cross"][c] for c in cs]
        for c,v in enumerate(vals,1):
            cc=s2.cell(row=rr,column=c,value=v); cc.font=base; cc.border=bd
            cc.alignment=C if c in cc2 else (Rg if c==13 else L)
            if c==13 and isinstance(v,(int,float)): cc.number_format='$#,##0'
            if rr%2==0: cc.fill=band
            if c>13: noflag(cc,v)
        rr+=1
    s2.auto_filter.ref=f"A3:{get_column_letter(n2)}{rr-1}"

    # Validation
    s3=wb.create_sheet("Validation"); s3.sheet_view.showGridLines=False
    s3["A1"]="Validation — 1 account per tier per rep, checked against actual orders"; s3["A1"].font=tf; s3.merge_cells("A1:G1")
    header(s3,3,["Rep","Tier","Location",f"{uw}s","# Orders","Order History (date · order# · qty)","Check"],[20,10,30,9,9,62,8]); rr=4
    for t in sample:
        hist=", ".join(f'{o["date"]} {o["name"]}={o["qty"]}' for o in t["hist"]); summ=sum(o["qty"] for o in t["hist"])
        vals=[t["rep"],t["tier"],t["location"],t["units"],len(t["hist"]),hist,"OK" if summ==t["units"] else f"!{summ}"]
        for c,v in enumerate(vals,1):
            cc=s3.cell(row=rr,column=c,value=v); cc.font=base; cc.border=bd
            cc.alignment=C if c in(2,4,5,7) else L
            if rr%2==0: cc.fill=band
            if c==7: cc.font=Font(name=FONT,bold=True,color=SAGE if v=="OK" else "B00000",size=10)
        rr+=1
    s3.auto_filter.ref=f"A3:G{rr-1}"
    wb.save(out)

if __name__=="__main__":
    import os
    csv_path=sys.argv[1] if len(sys.argv)>1 else None
    out_path=sys.argv[2] if len(sys.argv)>2 else "loyalty_report.xlsx"
    # territory map: arg3, else default file from assign_reps.mjs if present
    terr_path=sys.argv[3] if len(sys.argv)>3 else ("/tmp/territory_reps.json" if os.path.exists("/tmp/territory_reps.json") else None)
    if not csv_path:
        print("usage: loyalty_report.py <orders.json|export.csv> [output.xlsx] [territory_reps.json]"); sys.exit(1)
    t1,t2,s=build(csv_path,out_path,terr_path)
    tiers_n={}
    for t in t1: tiers_n[t["tier"]]=tiers_n.get(t["tier"],0)+1
    bad=[x for x in s if sum(o["qty"] for o in x["hist"])!=x["units"]]
    reassigned=sum(1 for t in t1 if t["openedBy"]); unassigned=sum(1 for t in t1 if t["rep"]=="(Unassigned)")
    print(f"OK  -> {out_path}   (territory: {'on' if terr_path else 'OFF — run assign_reps.mjs'})")
    print(f"  Tier Progress: {len(t1)} locations | tiers {tiers_n} | total units {sum(t['units'] for t in t1)}")
    print(f"  Rep: {reassigned} reassigned by territory | {unassigned} unassigned")
    print(f"  Lapsed: {len(t2)} | Validation sample: {len(s)} ({'all tie' if not bad else str(len(bad))+' MISMATCH'})")
