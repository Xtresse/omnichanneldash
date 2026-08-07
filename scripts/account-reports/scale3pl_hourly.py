#!/usr/bin/env python3
"""
Hourly Scale3PL order grid builder + Shopify reconciliation.

Pulls every UNFULFILLED order since the anchor date that is NOT tagged shipbob
(i.e. routes to Scale3PL, our west-coast 3PL), excluding Xvie and test orders,
and writes a filled Scale3PL EDI import grid plus a reconciliation log.

Ship Via:  DTC -> USPS,  everything else (B2B) -> UPSG.

Creds come from omnichanneldash/.env.local (SHOPIFY_CLIENT_ID/SECRET, client-
credentials grant) — same source as pull_orders.mjs. No MCP / no session needed.

Run:  python3 scale3pl_hourly.py
"""
import os, sys, json, urllib.request, datetime, pathlib
from copy import copy
import openpyxl

REPO   = pathlib.Path(__file__).resolve().parents[2]          # omnichanneldash/
ENV    = REPO / ".env.local"
TEMPLATE = pathlib.Path.home() / "Downloads" / "SAM XTR ORDER GRID 5-13-26.xlsx"
OUT_GRID = pathlib.Path.home() / "Downloads" / "Xtresse Order Uploads" / "SAM XTR ORDER GRID FILLED.xlsx"
OUT_RECON = REPO / "scripts" / "account-reports" / "scale3pl_reconciliation_live.csv"
LOG = REPO / "scripts" / "account-reports" / "scale3pl_hourly.log"
# Susie (Scale3PL) loads each file then wants the NEXT file to exclude what she already loaded.
LEDGER = REPO / "scripts" / "account-reports" / "scale3pl_sent_ledger.txt"

SINCE = os.environ.get("SCALE3PL_SINCE", "2026-06-23T18:00:00Z")   # 6/23 2pm EST
API   = os.environ.get("SHOPIFY_API_VERSION", "2025-01")

def load_env():
    if ENV.exists():
        for line in ENV.read_text().splitlines():
            m = line.strip()
            if "=" in m and not m.startswith("#"):
                k, v = m.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

def token(store):
    t = os.environ.get("SHOPIFY_ADMIN_API_TOKEN")
    if t: return t
    body = json.dumps({"client_id": os.environ["SHOPIFY_CLIENT_ID"],
                       "client_secret": os.environ["SHOPIFY_CLIENT_SECRET"],
                       "grant_type": "client_credentials"}).encode()
    req = urllib.request.Request(f"https://{store}/admin/oauth/access_token",
                                 data=body, headers={"Content-Type": "application/json"})
    return json.load(urllib.request.urlopen(req))["access_token"]

def gql(store, tok, query, variables=None):
    body = json.dumps({"query": query, "variables": variables or {}}).encode()
    req = urllib.request.Request(f"https://{store}/admin/api/{API}/graphql.json", data=body,
                                 headers={"Content-Type": "application/json", "X-Shopify-Access-Token": tok})
    return json.load(urllib.request.urlopen(req))

Q = """
query($q:String!,$after:String){
  orders(first:100, after:$after, query:$q, sortKey:CREATED_AT){
    edges{ node{
      name createdAt displayFinancialStatus tags
      shippingAddress{ name company address1 address2 city provinceCode zip countryCodeV2 phone }
      customer{ email phone }
      lineItems(first:50){ edges{ node{ quantity sku title } } }
    }}
    pageInfo{ hasNextPage endCursor }
  }
}"""

def fetch_orders():
    store = os.environ.get("SHOPIFY_STORE_DOMAIN", "ace1d0-26.myshopify.com")
    tok = token(store)
    q = (f"created_at:>='{SINCE}' fulfillment_status:unfulfilled -tag:SHIPBOB -tag:xvie")
    out, after = [], None
    while True:
        d = gql(store, tok, Q, {"q": q, "after": after})
        if "errors" in d: raise RuntimeError(d["errors"])
        conn = d["data"]["orders"]
        out += [e["node"] for e in conn["edges"]]
        if conn["pageInfo"]["hasNextPage"]: after = conn["pageInfo"]["endCursor"]
        else: break
    return out

def is_test(n):
    email = ((n.get("customer") or {}).get("email") or "").lower()
    return "samukhsood+" in email or n["displayFinancialStatus"] in ("VOIDED", "REFUNDED")

def channel(n):
    return "b2b" if "b2b" in [t.lower() for t in n["tags"]] else "dtc"

def build(orders):
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active
    tmpl = {c: copy(ws.cell(2, c)._style) for c in range(1, ws.max_column + 1)}
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    r = 2; rows = 0; flags = []
    for n in orders:
        ch = channel(n)
        sa = n.get("shippingAddress") or {}
        person = sa.get("name") or ""
        company = sa.get("company") or ""
        name = company if (ch == "b2b" and company) else person
        phone = sa.get("phone") or (n.get("customer") or {}).get("phone") or ""
        ship_via = "USPS" if ch == "dtc" else "UPSG"
        no = n["name"].lstrip("#")
        a1 = sa.get("address1") or ""
        if ch == "b2b" and a1.lower().startswith(("po box", "p.o", "pobox")):
            flags.append(f"{n['name']} B2B ships UPSG to PO box: {a1}")
        for li in n["lineItems"]["edges"]:
            it = li["node"]
            ws.cell(r,1).value=name; ws.cell(r,2).value=person; ws.cell(r,3).value=person
            ws.cell(r,4).value=a1; ws.cell(r,5).value=sa.get("address2") or ""
            ws.cell(r,6).value=sa.get("city") or ""; ws.cell(r,7).value=sa.get("provinceCode") or ""
            ws.cell(r,8).value=sa.get("zip") or ""; ws.cell(r,9).value=sa.get("countryCodeV2") or "US"
            ws.cell(r,10).value=(n.get("customer") or {}).get("email") or ""; ws.cell(r,11).value=phone
            ws.cell(r,23).value="TR"; ws.cell(r,24).value="XTR"
            ws.cell(r,25).value=("B2B" if ch=="b2b" else "DTC")
            ws.cell(r,26).value=str(no); ws.cell(r,34).value=ship_via
            ws.cell(r,40).value=it["sku"]; ws.cell(r,41).value=it["quantity"]; ws.cell(r,42).value=1
            for c in range(1, ws.max_column+1): ws.cell(r,c)._style = copy(tmpl[c])
            r += 1; rows += 1
    wb.save(OUT_GRID)
    return rows, flags

def recon(orders):
    with open(OUT_RECON, "w") as f:
        f.write("Order,Created (UTC),Channel,Ship Via,Financial,Ship To,Items\n")
        for n in orders:
            ch = channel(n)
            items = "; ".join(f"{li['node']['quantity']}x {li['node']['sku']}" for li in n["lineItems"]["edges"])
            sa = n.get("shippingAddress") or {}
            who = (sa.get("company") or sa.get("name") or "").replace(",", " ")
            f.write(f"{n['name']},{n['createdAt']},{ch},{'USPS' if ch=='dtc' else 'UPSG'},"
                    f"{n['displayFinancialStatus']},{who},{items}\n")

def main():
    load_env()
    stamp = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    mark = "--mark-sent" in sys.argv
    sent = {x.strip() for x in LEDGER.read_text().split() if x.strip()} if LEDGER.exists() else set()
    raw = fetch_orders()
    tests = [n["name"] for n in raw if is_test(n)]
    cur = [n for n in raw if not is_test(n)]
    # exclude orders Susie has already loaded (in the ledger). Shipped orders drop out automatically
    # because the Shopify query is fulfillment_status:unfulfilled.
    orders = [n for n in cur if n["name"].lstrip("#") not in sent]
    already = [n["name"] for n in cur if n["name"].lstrip("#") in sent]
    b2b = sum(1 for n in orders if channel(n) == "b2b")
    dtc = len(orders) - b2b
    if orders:
        rows, flags = build(orders)        # only overwrite the grid when there's something new
        recon(orders)
    else:
        rows, flags = 0, []
    if mark and orders:
        with open(LEDGER, "a") as f:
            for n in orders:
                f.write(n["name"].lstrip("#") + "\n")
    line = (f"[{stamp}] non-shipbob={len(cur)}  NEW_in_file={len(orders)} (B2B {b2b}/DTC {dtc})  "
            f"rows={rows}  already_loaded={len(already)}  marked_sent={'YES' if mark else 'no'}  "
            f"tests={tests}  flags={flags}")
    with open(LOG, "a") as f: f.write(line + "\n")
    print(line)
    print("Grid:", OUT_GRID)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        with open(LOG, "a") as f:
            f.write(f"[{datetime.datetime.now(datetime.timezone.utc):%Y-%m-%d %H:%M UTC}] ERROR {e}\n")
        print("ERROR:", e, file=sys.stderr); sys.exit(1)
