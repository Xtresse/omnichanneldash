#!/usr/bin/env python3
"""
Upload Scale3PL tracking (Susie's shipped-orders xlsx) to Shopify as fulfillments.

Default = DRY RUN (read-only): resolves each order, shows tracking + open fulfillment order.
  python3 upload_tracking.py <file.xlsx>
  python3 upload_tracking.py <file.xlsx> --apply [--limit N] [--no-notify]

Carrier map: tracking cell looks like "UPS2=Parcel=1ZA.." or "STAMPS=Parcel=9400..".
  prefix UPS*/UPS2 -> UPS ;  STAMPS -> USPS.  number = last '='-segment.
"""
import os, sys, json, urllib.request, pathlib
import openpyxl

REPO = pathlib.Path(__file__).resolve().parents[2]
ENV = REPO / ".env.local"
API = os.environ.get("SHOPIFY_API_VERSION", "2025-01")

def load_env():
    for line in ENV.read_text().splitlines():
        m = line.strip()
        if "=" in m and not m.startswith("#"):
            k, v = m.split("=", 1); os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

def token(store):
    if os.environ.get("SHOPIFY_ADMIN_API_TOKEN"): return os.environ["SHOPIFY_ADMIN_API_TOKEN"]
    body = json.dumps({"client_id": os.environ["SHOPIFY_CLIENT_ID"], "client_secret": os.environ["SHOPIFY_CLIENT_SECRET"], "grant_type": "client_credentials"}).encode()
    req = urllib.request.Request(f"https://{store}/admin/oauth/access_token", data=body, headers={"Content-Type": "application/json"})
    return json.load(urllib.request.urlopen(req))["access_token"]

STORE = None; TOK = None
def gql(q, v=None):
    body = json.dumps({"query": q, "variables": v or {}}).encode()
    req = urllib.request.Request(f"https://{STORE}/admin/api/{API}/graphql.json", data=body,
                                 headers={"Content-Type": "application/json", "X-Shopify-Access-Token": TOK})
    return json.load(urllib.request.urlopen(req))

def carrier_and_number(cell):
    cell = str(cell or "").strip()
    parts = cell.split("=")
    num = parts[-1].strip() if parts else ""
    pre = parts[0].strip().upper() if parts else ""
    comp = "UPS" if pre.startswith("UPS") else ("USPS" if pre.startswith("STAMP") else pre)
    return comp, num

def parse(path):
    ws = openpyxl.load_workbook(path, data_only=True).active
    hdr = {str(ws.cell(1,c).value).strip(): c for c in range(1, ws.max_column+1)}
    oc = hdr.get("Customer Order"); tc = hdr.get("Tracking Link(s)")
    rows = []
    for r in range(2, ws.max_row+1):
        o = ws.cell(r, oc).value
        if o in (None, ""): continue
        comp, num = carrier_and_number(ws.cell(r, tc).value)
        rows.append({"order": str(o).strip().lstrip("#"), "company": comp, "number": num,
                     "type": ws.cell(r, hdr.get(" Type", 2)).value})
    return rows

Q_ORDER = """query($q:String!){ orders(first:1, query:$q){ edges{ node{ id name displayFulfillmentStatus
  fulfillmentOrders(first:10){ edges{ node{ id status lineItems(first:50){ edges{ node{ id remainingQuantity } } } } } } } } } }"""

M_FULFILL = """mutation($f: FulfillmentV2Input!){ fulfillmentCreateV2(fulfillment:$f){
  fulfillment{ id status trackingInfo{ number company url } } userErrors{ field message } } }"""

def resolve(order_no):
    d = gql(Q_ORDER, {"q": f"name:{order_no}"})
    e = d.get("data", {}).get("orders", {}).get("edges", [])
    if not e: return None
    return e[0]["node"]

def main():
    global STORE, TOK
    if len(sys.argv) < 2: sys.exit("usage: upload_tracking.py <file.xlsx> [--apply] [--limit N] [--no-notify]")
    path = sys.argv[1]
    apply = "--apply" in sys.argv
    notify = "--no-notify" not in sys.argv
    limit = None
    if "--limit" in sys.argv: limit = int(sys.argv[sys.argv.index("--limit")+1])
    load_env(); STORE = os.environ.get("SHOPIFY_STORE_DOMAIN", "ace1d0-26.myshopify.com"); TOK = token(STORE)

    rows = parse(path)
    if "--emit" in sys.argv:
        out = []
        for row in rows:
            node = resolve(row["order"])
            if not node: continue
            if node["displayFulfillmentStatus"] == "FULFILLED": continue
            fos = [fo["node"]["id"] for fo in node["fulfillmentOrders"]["edges"]
                   if fo["node"]["status"] in ("OPEN","IN_PROGRESS","SCHEDULED")]
            if not fos: continue
            out.append({"order": row["order"], "fo": fos, "company": row["company"], "number": row["number"]})
        p = REPO / "scripts" / "account-reports" / "tofulfill.json"
        p.write_text(json.dumps(out, indent=0))
        print(f"emitted {len(out)} to-fulfill -> {p}")
        return
    print(f"Parsed {len(rows)} shipped rows. Mode: {'APPLY' if apply else 'DRY RUN'}  notify={notify}\n")
    done = ok = skip = err = 0
    for row in rows:
        if limit is not None and done >= limit: break
        done += 1
        node = resolve(row["order"])
        if not node:
            print(f"  #{row['order']:>5}  NOT FOUND in Shopify"); err += 1; continue
        fstatus = node["displayFulfillmentStatus"]
        open_fos = [fo["node"] for fo in node["fulfillmentOrders"]["edges"] if fo["node"]["status"] in ("OPEN","IN_PROGRESS","SCHEDULED")]
        tag = f"#{row['order']:>5}  {row['company']:>4} {row['number']:<24} [{fstatus}]"
        if fstatus == "FULFILLED" or not open_fos:
            print(f"  {tag}  already fulfilled -> skip"); skip += 1; continue
        if not apply:
            print(f"  {tag}  -> would fulfill ({len(open_fos)} FO)"); ok += 1; continue
        fo_inputs = [{"fulfillmentOrderId": fo["id"]} for fo in open_fos]
        var = {"f": {"lineItemsByFulfillmentOrder": fo_inputs, "notifyCustomer": notify,
                     "trackingInfo": {"number": row["number"], "company": row["company"]}}}
        res = gql(M_FULFILL, var)
        ue = res.get("data", {}).get("fulfillmentCreateV2", {}).get("userErrors") or res.get("errors")
        if ue:
            print(f"  {tag}  ERROR: {ue}"); err += 1
        else:
            f = res["data"]["fulfillmentCreateV2"]["fulfillment"]
            print(f"  {tag}  OK -> {f['status']}"); ok += 1
    print(f"\n{'Applied' if apply else 'Dry-run'}: {ok} {'fulfilled' if apply else 'to-fulfill'}, {skip} skipped, {err} errors, {done} processed")

if __name__ == "__main__":
    main()
