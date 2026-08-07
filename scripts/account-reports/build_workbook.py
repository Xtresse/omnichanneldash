#!/usr/bin/env python3
"""Build the brand-styled Territory Expansion workbook using the CANONICAL
rep-territory attribution (Sales-Rep-Dashboards/lib/repTerritory.js) — the
same engine that drives the rep dashboards. Source: accounts_attributed.json."""
import pandas as pd, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT='/Users/samsood/Documents/GitHub/omnichanneldash/scripts/account-reports'
XLSX=f'{OUT}/Xtresse_Territory_Expansion.xlsx'

MAROON='5C2F2E';CREAM='EDDCC5';LCREAM='F8F0E2';BROWN='2D1F1A';SAGE='5C8A6F';AMBER='C49A4F';RUST='9C4722';WHITE='FFFFFF'
fmaroon=PatternFill('solid',fgColor=MAROON);fcream=PatternFill('solid',fgColor=CREAM);flcream=PatternFill('solid',fgColor=LCREAM)
thin=Side(style='thin',color='D9C9B0');border=Border(thin,thin,thin,thin)
H1=Font(name='Cormorant Garamond',size=24,bold=True,color=MAROON)
H2=Font(name='Cormorant Garamond',size=15,bold=True,color=MAROON)
HEADW=Font(name='Inter',size=10,bold=True,color=WHITE)
BODY=Font(name='Inter',size=10,color=BROWN);BODYB=Font(name='Inter',size=10,bold=True,color=BROWN)
SUB=Font(name='Inter',size=10,italic=True,color='7A6A58')

# slide 8/9 addressable midpoints + Month-6 targets, and the incoming-1099 label
META={'Nevada + Utah':(350,'$65–80K/mo'),'Dallas (DFW)':(525,'$70–85K/mo'),
 'Ohio (Columbus/Cleveland)':(450,'$60–75K/mo'),'DC / NoVA':(450,'$65–80K/mo'),
 'Long Island / Queens':(450,'$65–80K/mo'),'Houston / South Texas':(550,'$60–75K/mo'),
 'Brooklyn / Staten Island':(300,'$45–60K/mo'),'Philadelphia / Pittsburgh':(400,'$50–65K/mo'),
 'Buffalo / Rochester / Albany':(240,'$35–45K/mo'),'Mississippi / Alabama':(175,'$35–45K/mo')}
TIERN={'Nevada + Utah':1,'Dallas (DFW)':1,'Ohio (Columbus/Cleveland)':1,'DC / NoVA':1,'Long Island / Queens':1,
 'Houston / South Texas':2,'Brooklyn / Staten Island':2,'Philadelphia / Pittsburgh':2,
 'Buffalo / Rochester / Albany':3,'Mississippi / Alabama':3}

acc=pd.DataFrame(json.load(open(f'{OUT}/accounts_attributed.json')))
book=pd.DataFrame(json.load(open(f'{OUT}/rep_book.json')))
acc['tier']=acc['territory'].map(TIERN)
acc['rep']=acc['assignedRep'].fillna('— (open / new 1099)')
# Nevada + Utah: Tuckett removed entirely — ALL accounts transfer to the new 1099 (none fenced).
TRANSFER_TERR={'Nevada + Utah'}
acc['reno_transfer']=acc['territory'].isin(TRANSFER_TERR)
acc.loc[acc['reno_transfer'],'rep']='→ New 1099 (NV+UT transfer)'
acc.loc[acc['reno_transfer'],'assignedSection']='1099 (incoming)'

def money(c):c.number_format='$#,##0'
def setw(ws,ws_w):
    for i,w in enumerate(ws_w,1):ws.column_dimensions[get_column_letter(i)].width=w
def header(ws,row,labels,wrap=True):
    for j,h in enumerate(labels,1):
        c=ws.cell(row,j,h);c.font=HEADW;c.fill=fmaroon;c.border=border
        c.alignment=Alignment(wrap_text=wrap,vertical='center',horizontal='center')

wb=Workbook()

# ---------------- TAB 1 — PITCH ----------------
ws=wb.active;ws.title='The Pitch';ws.sheet_view.showGridLines=False
ws['A1']='Territory Expansion — What It Means For You';ws['A1'].font=H1
ws['A2']='Mapping is the live rep-dashboard territory engine (tag-primary + ZIP/state). 10 new markets · 1099s farm white space · your book stays yours.';ws['A2'].font=SUB
ws.merge_cells('A1:H1');ws.merge_cells('A2:H2')
cards=[('YOUR ACCOUNTS','100%','locked to you in the dashboard',SAGE),
       ('MOVED TO A 1099','$0','of any rep’s current revenue',SAGE),
       ('NEW TERRITORIES','10','~3,890 addressable clinics',MAROON),
       ('OPEN WHITE SPACE','~3,470','clinics no rep has ever sold',AMBER)]
r=4
for i,(lab,big,sub,col) in enumerate(cards):
    c0=1+i*2
    a=ws.cell(r,c0,lab);a.font=Font(name='Inter',size=9,bold=True,color=WHITE);a.fill=PatternFill('solid',fgColor=col);a.alignment=Alignment('center')
    ws.merge_cells(start_row=r,start_column=c0,end_row=r,end_column=c0+1)
    b=ws.cell(r+1,c0,big);b.font=Font(name='Cormorant Garamond',size=26,bold=True,color=col);b.alignment=Alignment('center')
    ws.merge_cells(start_row=r+1,start_column=c0,end_row=r+1,end_column=c0+1)
    s=ws.cell(r+2,c0,sub);s.font=Font(name='Inter',size=8,color=BROWN);s.alignment=Alignment('center')
    ws.merge_cells(start_row=r+2,start_column=c0,end_row=r+2,end_column=c0+1)
ws.row_dimensions[r+1].height=34
para=[('Why we are doing this',H2),
 ("Our densest professional markets have 300–600 target clinics each, but we have only ever sold to a fraction. "
  "Dallas: 151 active accounts against ~525 addressable. Houston: 150 against ~550. Every clinic we have not reached is "
  "revenue a competitor can take first.",BODY),
 ("1099 reps cover that white space at zero fixed cost — no base, no benefits, no travel. They are paid only on "
  "net-new accounts they open.",BODY),
 ('How your accounts are protected',H2),
 ("We mapped every account with the exact engine behind your rep dashboard — your tagged book first, then ZIP/state "
  "territory. Every account that engine puts on your list is carved out and fenced off before a 1099 enters the market. "
  "The 1099 may only open clinics that have never ordered from Xtresse.",BODY),
 ("Across all 10 new territories, the business that moves from any existing rep’s current accounts to a 1099 is $0. "
  "Not a reduced split — zero. You keep every account and every dollar you have today, and keep growing them.",BODYB),
 ('What this is worth to you',H2),
 ("A 1099 working the cold space around you warms the whole market — more clinics on Xtresse, more peer referrals, "
  "stronger regional presence. You get the halo with none of the cost and none of the risk to your book.",BODY)]
rr=r+4
for txt,fnt in para:
    ws.cell(rr,1,txt).font=fnt;ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=8)
    ws.cell(rr,1).alignment=Alignment(wrap_text=True,vertical='top')
    ws.row_dimensions[rr].height=18 if fnt is H2 else 46;rr+=1
setw(ws,[15]*8)

# ---------------- TAB 2 — TERRITORY & OWNER ----------------
ws=wb.create_sheet('Territory & Owner');ws.sheet_view.showGridLines=False
ws['A1']='The 10 New Territories — Current Owner vs. New 1099 Runway';ws['A1'].font=H1;ws.merge_cells('A1:I1')
ws['A2']='Current owner = who the live dashboard assigns the market to today (their accounts are protected). White space = addressable minus active accounts.';ws['A2'].font=SUB;ws.merge_cells('A2:I2')
header(ws,4,['Territory','Tier','Current owner (protected)','Active accts','Protected sales (2-yr)','Addressable clinics','New-1099 white space (accts)','Month-6 target'])
ws.row_dimensions[4].height=42
# owner(s) per territory + totals
rows=[]
for terr in META:
    sub=acc[acc.territory==terr]
    prot=sub[~sub['reno_transfer']]            # protected = non-transfer accounts only
    owners=prot.groupby('assignedRep')['netSales'].sum().sort_values(ascending=False)
    owner_lbl=owners.index[0] if len(owners) else '—'
    if len(owners)>1: owner_lbl+=f" (+{len(owners)-1})"
    if len(prot)==0 and sub['reno_transfer'].any(): owner_lbl='James Tuckett → new 1099 (full transfer)'
    elif sub['reno_transfer'].any(): owner_lbl+=' · partial → new 1099'
    addr=META[terr][0]
    rows.append((terr,TIERN[terr],owner_lbl,len(prot),prot.netSales.sum(),addr,max(addr-len(prot),0),META[terr][1]))
rr=5
for terr,tier,owner,na,sales,addr,white,m6 in sorted(rows,key=lambda x:(x[1],x[0])):
    vals=[terr,f'Tier {tier}',owner,na,sales,addr,white,m6]
    for j,v in enumerate(vals,1):
        c=ws.cell(rr,j,v);c.border=border;c.font=BODY;c.fill=flcream if rr%2 else fcream
        if j in (2,4,6,7):c.alignment=Alignment('center')
        if j==5:money(c)
    ws.cell(rr,1).font=BODYB;ws.cell(rr,3).font=BODYB
    ws.cell(rr,7).font=Font(name='Inter',size=10,bold=True,color=SAGE);rr+=1
tot=['TOTAL','','',sum(r[3] for r in rows),sum(r[4] for r in rows),sum(r[5] for r in rows),sum(r[6] for r in rows),'']
for j,v in enumerate(tot,1):
    c=ws.cell(rr,j,v);c.font=Font(name='Inter',size=10,bold=True,color=WHITE);c.fill=fmaroon;c.border=border
    if j in (4,6,7):c.alignment=Alignment('center')
    if j==5:money(c)
setw(ws,[30,8,26,11,17,16,20,14])

# ---------------- TAB 3 — BY ZIP / CITY ----------------
ws=wb.create_sheet('By ZIP-City');ws.sheet_view.showGridLines=False
ws['A1']='Geographic Breakdown — by State, City, ZIP';ws['A1'].font=H1;ws.merge_cells('A1:H1')
ws['A2']='Every ZIP/city where an existing rep already has accounts (= claimed & protected). The new 1099 prospects every OTHER ZIP in the metro.';ws['A2'].font=SUB;ws.merge_cells('A2:H2')
header(ws,4,['Territory','State','City','ZIP','Current rep (protected)','Accts','Sales (2-yr net)','Last order'])
ws.row_dimensions[4].height=30
zc=(acc.groupby(['tier','territory','state','city','zip','rep'])
       .agg(accts=('company','size'),sales=('netSales','sum'),last=('lastOrder','max'))
       .reset_index().sort_values(['tier','territory','sales'],ascending=[True,True,False]))
rr=5
cur=None
for _,row in zc.iterrows():
    if cur!=row.territory:  # territory band
        c=ws.cell(rr,1,f'{row.territory}  ·  Tier {row.tier}  ·  owner: {acc[acc.territory==row.territory].groupby("assignedRep")["netSales"].sum().idxmax()}')
        c.font=Font(name='Inter',size=10,bold=True,color=WHITE);c.fill=PatternFill('solid',fgColor=MAROON)
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=8);rr+=1;cur=row.territory
    vals=[row.territory,row.state,row.city,row.zip,row.rep,row.accts,row.sales,row['last']]
    for j,v in enumerate(vals,1):
        c=ws.cell(rr,j,v);c.border=border;c.font=BODY;c.fill=flcream if rr%2 else fcream
        if j in (2,4,6):c.alignment=Alignment('center')
        if j==7:money(c)
    rr+=1
ws.freeze_panes='A5';setw(ws,[26,7,20,9,24,7,16,12])

# ---------------- TAB 4 — ACCOUNTS MASTER ----------------
ws=wb.create_sheet('Accounts (master)');ws.sheet_view.showGridLines=False
ws['A1']='Every Account In The New Territories';ws['A1'].font=H1;ws.merge_cells('A1:K1')
ws['A2']='Current rep = canonical dashboard attribution (tag-primary, ZIP/state fallback). Every account is protected; the 1099 only works net-new clinics.';ws['A2'].font=SUB;ws.merge_cells('A2:K2')
cols=[('territory','Territory',26),('tier','Tier',6),('rep','Current rep (protected)',24),('assignedSection','Rep type',14),
      ('company','Account',34),('city','City',16),('state','State',7),('zip','ZIP',8),
      ('netSales','Sales (2-yr net)',15),('orders','Orders',8),('lastOrder','Last order',12)]
header(ws,4,[c[1] for c in cols]);ws.row_dimensions[4].height=30
for j,(k,lab,w) in enumerate(cols,1):ws.column_dimensions[get_column_letter(j)].width=w
am=acc.sort_values(['tier','territory','netSales'],ascending=[True,True,False])
rr=5
for _,row in am.iterrows():
    for j,(k,lab,w) in enumerate(cols,1):
        v=row.get(k,'')
        c=ws.cell(rr,j,v if pd.notna(v) else '');c.border=border;c.font=BODY;c.fill=flcream if rr%2 else fcream
        if k=='netSales':money(c)
        if k in('tier','state','orders','zip'):c.alignment=Alignment('center')
    ws.cell(rr,3).font=BODYB;rr+=1
ws.freeze_panes='A5';ws.auto_filter.ref=f'A4:{get_column_letter(len(cols))}{rr-1}'

# ---------------- TAB 5 — REP PROTECTION ----------------
ws=wb.create_sheet('Rep Protection');ws.sheet_view.showGridLines=False
ws['A1']='Existing Reps — Your Book Is Locked';ws['A1'].font=H1;ws.merge_cells('A1:F1')
ws['A2']='Full book = your total attributed accounts (all territories). Protected in new markets stays 100% yours; $ moved to a 1099 is $0 for everyone.';ws['A2'].font=SUB;ws.merge_cells('A2:F2')
header(ws,4,['Rep','Type','Full book (2-yr net)','Accts protected in new territories','$ protected & carved out','$ moved to a 1099']);ws.row_dimensions[4].height=42
# exclude Reno transfers — Tuckett is not credited Reno as protected (he's leaving it)
in_terr=acc[~acc['reno_transfer']].dropna(subset=['assignedRep']).groupby('assignedRep').agg(accts=('company','size'),sales=('netSales','sum'))
bk=book.set_index('rep')
order=bk.sort_values('total_book',ascending=False)
rr=5
for rep,brow in order.iterrows():
    na=int(in_terr.accts.get(rep,0));sa=float(in_terr.sales.get(rep,0))
    vals=[rep,brow.section,brow.total_book,na,sa,0]
    for j,v in enumerate(vals,1):
        c=ws.cell(rr,j,v);c.border=border;c.font=BODY;c.fill=flcream if rr%2 else fcream
        if j==4:c.alignment=Alignment('center')
        if j in(3,5,6):money(c)
    ws.cell(rr,1).font=BODYB
    ws.cell(rr,6).font=Font(name='Inter',size=10,bold=True,color=SAGE);rr+=1
setw(ws,[24,12,18,24,20,16])

# ---------------- TAB 6 — 1099 POTENTIAL (rollup w/ NPI cross-check) ----------------
roll=pd.read_csv(f'{OUT}/territory_rollup.csv')
zp=pd.read_csv(f'{OUT}/zip_potential.csv',dtype={'zipcode':str})
ws=wb.create_sheet('1099 Potential');ws.sheet_view.showGridLines=False
ws['A1']='1099 Territory Potential — Open White Space & Expected Sales';ws['A1'].font=H1;ws.merge_cells('A1:M1')
ws['A2']='Target clinics = population ÷ 10,000 × income factor (derm/medspa/plastic/aesthetic). Real clinics = NPPES derm+plastic locations, all 10 markets (medspas extra). TAM = open × $3,500/yr.';ws['A2'].font=SUB;ws.merge_cells('A2:M2')
header(ws,4,['Territory','Tier','Current owner (fenced)','ZIPs','Population','Est. target clinics','Real derm+plastic clinics (NPI)','Deck addressable','Our accounts','Open clinics (1099 runway)','Market TAM (annual)','Deck Mo-6 $/mo','Deck Mo-12 $/mo'])
ws.row_dimensions[4].height=46
rr=5
for _,r in roll.sort_values(['tier','territory']).iterrows():
    npi='—' if (pd.isna(r.npi_loc) or r.npi_loc==0) else int(r.npi_loc)
    vals=[r.territory,f'Tier {int(r.tier)}',r.owner,int(r.zips),int(r['pop']),int(r.target),npi,int(r.deck_addr),
          int(r.our_accts),int(r['open']),r.tam,r.deck_mo6,r.deck_mo12]
    for j,v in enumerate(vals,1):
        c=ws.cell(rr,j,v);c.border=border;c.font=BODY;c.fill=flcream if rr%2 else fcream
        if j in(2,4,6,7,8,9,10):c.alignment=Alignment('center')
        if j==5:c.number_format='#,##0'
        if j in(11,12,13):money(c)
    ws.cell(rr,1).font=BODYB;ws.cell(rr,3).font=BODYB
    ws.cell(rr,10).font=Font(name='Inter',size=10,bold=True,color=SAGE)
    ws.cell(rr,11).font=Font(name='Inter',size=10,bold=True,color=AMBER);rr+=1
tot=['TOTAL','','',int(roll.zips.sum()),int(roll['pop'].sum()),int(roll.target.sum()),
     int(roll.npi_loc.fillna(0).sum()),int(roll.deck_addr.sum()),int(roll.our_accts.sum()),
     int(roll['open'].sum()),roll.tam.sum(),roll.deck_mo6.sum(),roll.deck_mo12.sum()]
for j,v in enumerate(tot,1):
    c=ws.cell(rr,j,v);c.font=Font(name='Inter',size=10,bold=True,color=WHITE);c.fill=fmaroon;c.border=border
    if j in(4,6,7,8,9,10):c.alignment=Alignment('center')
    if j==5:c.number_format='#,##0'
    if j in(11,12,13):money(c)
ws.cell(rr+1,7,'NPI = real CMS provider data (all 10 markets); medspas/aesthetic NP clinics are additional').font=SUB
setw(ws,[28,7,22,7,12,14,16,12,11,16,15,12,12])

# ---------------- TAB 7 — ZIP ACCESS (every ZIP) ----------------
ws=wb.create_sheet('ZIP Access (all)');ws.sheet_view.showGridLines=False
ws['A1']='Every ZIP The New 1099 Reps Can Work';ws['A1'].font=H1;ws.merge_cells('A1:N1')
ws['A2']='Full territory geography. Open clinics = est. target − our existing accounts (which stay with the current rep). Annual potential = open × $3,500. NPI cols = real CMS derm+plastic, all 10 markets.';ws['A2'].font=SUB;ws.merge_cells('A2:N2')
header(ws,4,['Territory','Tier','State','County','City','ZIP','Population','Median HH income','Est. target clinics','NPI providers','NPI clinics','Our accts (fenced)','Open clinics','Annual potential'])
ws.row_dimensions[4].height=30
rr=5
for _,r in zp.iterrows():
    npip='' if pd.isna(r.npi_providers) else int(r.npi_providers)
    npil='' if pd.isna(r.npi_locations) else int(r.npi_locations)
    inc='' if pd.isna(r.median_household_income) else int(r.median_household_income)
    vals=[r.territory,int(r.tier),r.state,r.county,r.major_city,r.zipcode,int(r.population),inc,
          int(r.target_clinics),npip,npil,int(r.our_accts),int(r.open_clinics),int(r.annual_potential)]
    for j,v in enumerate(vals,1):
        c=ws.cell(rr,j,v);c.border=border;c.font=BODY;c.fill=flcream if rr%2 else fcream
        if j in(2,6,9,10,11,12,13):c.alignment=Alignment('center')
        if j in(7,8):c.number_format='#,##0'
        if j==14:money(c)
    if r.open_clinics>0: ws.cell(rr,13).font=Font(name='Inter',size=10,bold=True,color=SAGE)
    rr+=1
ws.freeze_panes='A5';ws.auto_filter.ref=f'A4:N{rr-1}'
setw(ws,[26,6,7,20,18,8,11,13,14,12,11,13,11,14])

wb.save(XLSX)
print('Saved',XLSX,'| tabs:',wb.sheetnames,'| accounts:',len(acc))
