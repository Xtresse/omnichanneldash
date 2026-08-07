#!/usr/bin/env python3
"""Fill the Scale3PL EDI order grid from the 46 non-shipbob (excl Xvie/test) orders."""
import openpyxl
from copy import copy

SRC = "/Users/samsood/Downloads/SAM XTR ORDER GRID 5-13-26.xlsx"
OUT = "/Users/samsood/Downloads/SAM XTR ORDER GRID 5-13-26 FILLED.xlsx"

# order_no, channel, person, company, addr1, addr2, city, state, zip, phone, email, [(sku,qty),...]
O = [
("9072","dtc","Amber Withycombe","","11204 Cap Stone Drive","","Austin","TX","78739","","amberwithy@gmail.com",[("X-GN-060CT-001",3)]),
("9073","b2b","Abby Lovell","Honeybee NP Aesthetics & Wellness","3744 County Road 223","","Stephenville","TX","76401","+12544858298","abbylovellnp@yahoo.com",[("860011740100",1)]),
("9076","dtc","Christina V","","1516 Lorraine Dr","","Plano","TX","75074","","brogan.50-conceit@icloud.com",[("X-GN-060CT-001",1)]),
("9078","dtc","Gabrielle Sharaga","","1025 Coronado Terrace","","Los Angeles","CA","90026","+18057945174","gsharaga@gmail.com",[("X-GN-060CT-001",3)]),
("9083","dtc","Denise Friou","","2615 Drew Street","","Houston","TX","77004","+17135607242","dfriou@email.com",[("X-GN-060CT-001",1)]),
("9084","dtc","Lisa Sliger","","1295 Big Horn Ct","","Middleton","ID","83644","","lisasliger1978@gmail.com",[("X-GN-060CT-001",1)]),
("9086","b2b","Meredith Abbott","Abbott Aesthetics","4702 West Lovers Lane","","Dallas","TX","75209","+12149450466","meredith@doctorabbott.com",[("860011740100",1)]),
("9087","b2b","Irina Minina","Skin Perfections","1201 E Alton Gloor Blvd","","Brownsville","TX","78526","+19565422188","iraminina@aol.com",[("860011740100",1)]),
("9091","dtc","Shelly Brown","","Po box 213","","Gardendale","TX","79758","+14326383749","shellylbrown215@gmail.com",[("X-GN-060CT-001",1)]),
("9093","dtc","Bridget Marburger","","279 Wellborn Rd","","Liberty Hill","TX","78642-4948","+19794212211","bridget.bilski@gmail.com",[("X-GN-060CT-001",1)]),
("9094","dtc","Laurice Richardson","","7212 Sandy Ridge Rd","","Pasco","WA","99301","","vane.coppery.5c@icloud.com",[("X-GN-060CT-001",3)]),
("9095","dtc","Tricia OBrien","","10702 Beechwood Drive","","Rancho Cucamonga","CA","91737","+16262414141","triciaobrien4@yahoo.com",[("X-GN-060CT-001",1)]),
("9096","b2b","Shay Sim","Kucumber Skin Lounge","2001 1st Ave","","Seattle","WA","98121","+12064419300","chris@kucumberskinlounge.com",[("860011740100",1),("X-FRC-30ML-CASE",1),("X-GN-002CT-001",1)]),
("9097","b2b","Shay Sim","Kucumber Skin Lounge","10708 Main St","Suite 310","Bellevue","WA","98004","+14254293694","chris@kucumberskinlounge.com",[("860011740100",1),("X-GN-002CT-001",1)]),
("9098","dtc","Rebecca Curry","","1389 Wyoming Street","","Golden","CO","80403","3038070924","becky.curry@xtresse.com",[("X-GN-060CT-001",1)]),
("9100","b2b","Andrea Smith","SmithCare Dermatology","21800 Market Place Northwest","","Poulsbo","WA","98370","+13602915700","andreadsmithmd@gmail.com",[("860011740100",1)]),
("9101","dtc","Lori Severidt","","11872 W Asbury Pl","","Lakewood","CO","80228","","alanah@danieltoft.co",[("X-GN-060CT-001",1)]),
("9102","dtc","nancy ortiz","","1961 County Road 2021","","Glen Rose","TX","76043","","nortiz03@live.com",[("X-GN-060CT-001",1)]),
("9104","dtc","cynthia reader","","16771 Loma Street","","Los Gatos","CA","95032","","cynalicious@yahoo.com",[("X-GN-060CT-001",1)]),
("9106","b2b","Richmond Ramirez","Kutis Medical Aesthetics","2405 W Horizon Ridge Pkwy","100","Henderson","NV","89052","+17027793790","info@inspiradapc.com",[("X-GN-002CT-001",1),("860011740100",1)]),
("9109","dtc","Carol Goldberg","","1314 Falcon Ledge Drive","Apt 116","Austin","TX","78746","","carolmgoldberg@gmail.com",[("X-GN-060CT-001",1)]),
("9111","dtc","Amelia Genao","","200 North Bishop Avenue","458","Dallas","TX","75208","+19788264151","amelia.genao@yahoo.com",[("X-GN-060CT-001",1)]),
("9112","dtc","Becky Cohen","","921 Eolus Avenue","","Encinitas","CA","92024-2142","","beckycohenphotographer@gmail.com",[("X-GN-060CT-001",3)]),
("9114","dtc","Diane Swanson","","10211 Prestwick Trail","","Lone Tree","CO","80124","+13038813133","Dianeswanson303@gmail.com",[("XTR-DTC-GMFR-02",1)]),
("9119","dtc","Bonnie Beebe","","2794 North Sunrise Way","","Palm Springs","CA","92262","+17608353365","bonniebb465@gmail.com",[("X-GN-060CT-001",2)]),
("9121","dtc","Jackie Haines","","193 Mammoth Fork Drive","","Bozeman","MT","59718","+14065999898","jackiej.haines@gmail.com",[("X-GN-060CT-001",1)]),
("9123","dtc","David Jackson","","5672 N Mina Vis","","Tucson","AZ","85718-4122","","davidja10@aol.com",[("X-GN-060CT-001",2)]),
("9126","b2b","Michelle Aszterbaum","The Dermatology Center of Newport","360 San Miguel Dr.","Suite 406","Newport Beach","CA","92660","+19495250700","draszterbaum@gmail.com",[("860011740100",1)]),
("9127","b2b","Jamie Eschete","Desert Bloom Aesthetics","5519 W Lariat Ln","","Phoenix","AZ","85083","+16027036265","jamiern@cox.net",[("X-FRC-30ML-CASE",1)]),
("9129","b2b","Carly Stillman","Luma Medical Aesthetics","200 Market St","","Basalt","CO","81621","+19709899226","carlys@lumamedicalaesthetics.co",[("860011740100",1),("X-GN-002CT-001",1)]),
("9132","b2b","Melynda Fenn","AZ Vitality and Wellness","7984 West Happy Valley Road","","Peoria","AZ","85383","+16234141994","mfenn@azivvitality.com",[("860011740100",2),("X-FRC-30ML-CASE",1)]),
("9135","dtc","Jeanna Mane","","123 Cedar St","","Lewistown","MT","59457","","jeannajmane@gmail.com",[("X-GN-060CT-001",3)]),
("9136","b2b","Ernai Hernandez","Mederna Aesthetics and Wellness-Attn: Hilda Banegas","5225 Katy Fwy","Suite 370","Houston","TX","77007","+18325248438","ernaihernandez@outlook.com",[("860011740100",1)]),
("9140","dtc","Caitlin Liang","","3355 Overland Avenue","506","Los Angeles","CA","90034","+13107514947","caitlin.liang@gmail.com",[("X-GN-060CT-001",3)]),
("9142","dtc","Karen Lashlee","","3969 Adams Street Apt E118","E118 1st Floor","Carlsbad","CA","92008","7609780409","klash.lashlee@gmail.com",[("X-GN-060CT-001",2)]),
("9144","b2b","Tricia Dikes","Gerald Minniti, MD, F.A.C.S.","15319 Del Gado Dr","","Sherman Oaks","CA","91403","+13102750040","triciadikes@hotmail.com",[("860011740100",1)]),
("9150","dtc","Uzma Zafar","","3111 Turtle Head Peak Dr","","Las Vegas","NV","89135","+17023011115","uzikayani@yahoo.com",[("X-GN-060CT-001",1)]),
("9151","dtc","Becky Blum","","2091 W Dry Creek Rd","","Littleton","CO","80120","","beckyblum22@gmail.com",[("X-FRC-30ML-001",1),("X-GN-060CT-001",1)]),
("9154","dtc","Aaron Rogers","","1507 Northeast 12th Place","","Canby","OR","97013","+19499102866","arog2436@yahoo.com",[("X-GN-060CT-001",1)]),
("9167","dtc","Joanna Manning","","2510 N Henderson Ave","Apt 01","Dallas","TX","75206","","manningfamily7@aol.com",[("X-GN-060CT-001",1)]),
("9188","b2b","Pam Tate","OC Dermatology (LGN)","30201 Golden Lantern","Suite B","Laguna Niguel","CA","92677","+19493631788","ap@platinumderm.com",[("860011740100",3),("X-FRC-30ML-CASE",2)]),
("9192","b2b","Stormie Tuma","Beautiful Aesthetics","400 Stonebrook Parkway","Ste 104","Frisco","TX","75036","+12146180264","storm@beautifulaesthetics.com",[("860011740100",1)]),
("9194","b2b","Jessica Jackson","Restoration Haus","101 S Trinity St","","Decatur","TX","76234","+19406269097","restorationhaustx@gmail.com",[("860011740100",1),("X-GN-002CT-002",1)]),
("9196","b2b","Jessica Jackson","Restoration Haus","101 S Trinity St","","Decatur","TX","76234","+19406269097","restorationhaustx@gmail.com",[("X-GN-002CT-001",1)]),
("9200","b2b","Sandi Eivins","Aesthetica","940 Central Park Drive","Suite 207","Steamboat Springs","CO","80487","+19708464730","drsandi@steamboatderm.com",[("860011740100",1),("X-GN-002CT-001",1)]),
("9209","dtc","Jennifer Carrillo","","1755 Lewis Pl NW","","Bainbridge Island","WA","98110","+12067944257","jencarrillo@gmail.com",[("X-GN-060CT-001",3)]),
]

wb = openpyxl.load_workbook(SRC)
ws = wb.active

# capture style of an existing data row (row 2) to mirror formatting
tmpl = {c: copy(ws.cell(row=2, column=c)._style) for c in range(1, ws.max_column+1)}

# clear all existing data rows (2..max)
for r in range(2, ws.max_row+1):
    for c in range(1, ws.max_column+1):
        ws.cell(row=r, column=c).value = None

def setrow(r, o, sku, qty):
    no, chan, person, company, a1, a2, city, st, zp, phone, email = o[:11]
    name = company if company else person
    vals = {
        1:name, 2:person, 3:person, 4:a1, 5:a2, 6:city, 7:st, 8:zp, 9:"US",
        10:email, 11:phone,
        23:"TR", 24:"XTR", 25:("B2B" if chan=="b2b" else "DTC"),
        26:"UPSG",                # AH Ship Via
        34:str(no),               # Z? -> actually Z is col 26... fix below
    }
    # explicit columns by letter to avoid index confusion
    ws.cell(r,1).value=name; ws.cell(r,2).value=person; ws.cell(r,3).value=person
    ws.cell(r,4).value=a1; ws.cell(r,5).value=a2; ws.cell(r,6).value=city
    ws.cell(r,7).value=st; ws.cell(r,8).value=zp; ws.cell(r,9).value="US"
    ws.cell(r,10).value=email; ws.cell(r,11).value=phone
    ws.cell(r,23).value="TR"      # W Business Unit
    ws.cell(r,24).value="XTR"     # X Trading Partner
    ws.cell(r,25).value=("B2B" if chan=="b2b" else "DTC")  # Y Order Type
    ws.cell(r,26).value=str(no)   # Z Order Number
    ws.cell(r,34).value="UPSG"    # AH Ship Via
    ws.cell(r,40).value=sku       # AN Item
    ws.cell(r,41).value=qty       # AO Requested Quantity
    ws.cell(r,42).value=1         # AP Unit Price
    for c in range(1, ws.max_column+1):
        ws.cell(r,c)._style = copy(tmpl[c])

r = 2
lines = 0
for o in O:
    for sku, qty in o[11]:
        setrow(r, o, sku, qty)
        r += 1; lines += 1

wb.save(OUT)
print(f"Orders: {len(O)}  Line-item rows written: {lines}  (rows 2..{r-1})")
print("Saved:", OUT)
