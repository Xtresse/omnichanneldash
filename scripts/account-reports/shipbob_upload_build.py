#!/usr/bin/env python3
"""
Build the ShipBob bulk order-upload file for orders that route to ShipBob
(tag:SHIPBOB, excl Xvie, unfulfilled since anchor, excl test orders).

DTC -> "Add B2C Orders Here" tab,  B2B -> "Add B2B Orders Here" tab.
Items use the raw Shopify SKU; StoreOrderId = Shopify order # (lets ShipBob
dedupe so re-uploads don't double-create).

Same live Shopify pull as scale3pl_hourly.py (client-credentials, .env.local).
Run: python3 shipbob_upload_build.py
"""
import os, json, urllib.request, datetime, pathlib
import openpyxl

REPO = pathlib.Path(__file__).resolve().parents[2]
ENV  = REPO / ".env.local"
TEMPLATE = pathlib.Path.home() / "Downloads" / "OrderImportTemplate11.2019.xlsx"
OUT = pathlib.Path.home() / "Downloads" / "SHIPBOB Bulk Upload FILLED.xlsx"
LOG = REPO / "scripts" / "account-reports" / "scale3pl_hourly.log"
SINCE = os.environ.get("SCALE3PL_SINCE", "2026-06-23T18:00:00Z")
API = os.environ.get("SHOPIFY_API_VERSION", "2025-01")

def load_env():
    if ENV.exists():
        for line in ENV.read_text().splitlines():
            m = line.strip()
            if "=" in m and not m.startswith("#"):
                k, v = m.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

def token(store):
    if os.environ.get("SHOPIFY_ADMIN_API_TOKEN"): return os.environ["SHOPIFY_ADMIN_API_TOKEN"]
    body = json.dumps({"client_id": os.environ["SHOPIFY_CLIENT_ID"], "client_secret": os.environ["SHOPIFY_CLIENT_SECRET"], "grant_type": "client_credentials"}).encode()
    req = urllib.request.Request(f"https://{store}/admin/oauth/access_token", data=body, headers={"Content-Type": "application/json"})
    return json.load(urllib.request.urlopen(req))["access_token"]

Q = """
query($q:String!,$after:String){ orders(first:100, after:$after, query:$q, sortKey:CREATED_AT){
  edges{ node{ name displayFinancialStatus displayFulfillmentStatus tags
    shippingAddress{ name company address1 address2 city provinceCode zip countryCodeV2 phone }
    customer{ email phone }
    lineItems(first:50){ edges{ node{ quantity sku } } } } }
  pageInfo{ hasNextPage endCursor } } }"""

def fetch():
    store = os.environ.get("SHOPIFY_STORE_DOMAIN", "ace1d0-26.myshopify.com")
    tok = token(store)
    q = f"created_at:>='{SINCE}' fulfillment_status:unfulfilled tag:SHIPBOB -tag:xvie"
    out, after = [], None
    while True:
        body = json.dumps({"query": Q, "variables": {"q": q, "after": after}}).encode()
        req = urllib.request.Request(f"https://{store}/admin/api/{API}/graphql.json", data=body, headers={"Content-Type": "application/json", "X-Shopify-Access-Token": tok})
        d = json.load(urllib.request.urlopen(req))
        conn = d["data"]["orders"]; out += [e["node"] for e in conn["edges"]]
        if conn["pageInfo"]["hasNextPage"]: after = conn["pageInfo"]["endCursor"]
        else: break
    return out

def is_test(n):
    email = ((n.get("customer") or {}).get("email") or "").lower()
    return "samukhsood+" in email or n["displayFinancialStatus"] in ("VOIDED", "REFUNDED")

def channel(n):
    return "b2b" if "b2b" in [t.lower() for t in n["tags"]] else "dtc"

def main():
    load_env()
    raw = fetch()
    # strictly UNFULFILLED only — IN_PROGRESS / PARTIALLY_FULFILLED are already in ShipBob
    orders = [n for n in raw if not is_test(n) and n.get("displayFulfillmentStatus") == "UNFULFILLED"]
    skipped = [n["name"] for n in raw if not is_test(n) and n.get("displayFulfillmentStatus") != "UNFULFILLED"]
    wb = openpyxl.load_workbook(TEMPLATE)
    b2c, b2b = wb["Add B2C Orders Here"], wb["Add B2B Orders Here"]
    rc, rb = 2, 2; n_c = n_b = 0
    for n in orders:
        sa = n.get("shippingAddress") or {}
        person = sa.get("name") or ""; company = sa.get("company") or ""
        phone = sa.get("phone") or (n.get("customer") or {}).get("phone") or ""
        email = (n.get("customer") or {}).get("email") or ""
        items = [(li["node"]["sku"], li["node"]["quantity"]) for li in n["lineItems"]["edges"]]
        no = n["name"]
        if channel(n) == "dtc":
            ws, r = b2c, rc
            base = [person, sa.get("address1") or "", sa.get("address2") or "", sa.get("city") or "",
                    sa.get("provinceCode") or "", sa.get("zip") or "", sa.get("countryCodeV2") or "US",
                    email, phone, no, "", ""]          # ...StoreOrderId, ExtraInfo, ItemInfo
            for c, v in enumerate(base, start=1): ws.cell(r, c).value = v
            col = 13                                    # Item1
            for sku, qty in items:
                ws.cell(r, col).value = sku; ws.cell(r, col+1).value = qty; col += 2
            rc += 1; n_c += 1
        else:
            ws, r = b2b, rb
            base = [company or person, sa.get("address1") or "", sa.get("address2") or "", sa.get("city") or "",
                    sa.get("provinceCode") or "", sa.get("zip") or "", sa.get("countryCodeV2") or "US",
                    email, phone, "", "", "", ""]       # ...B2BLabelType, packing, sku-instr, reserve-date
            for c, v in enumerate(base, start=1): ws.cell(r, c).value = v
            col = 14                                    # Item1
            for sku, qty in items:
                ws.cell(r, col).value = sku; ws.cell(r, col+1).value = qty; col += 2
            rb += 1; n_b += 1
    wb.save(OUT)
    stamp = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    line = f"[{stamp}] ShipBob upload built: {len(orders)} orders (B2C {n_c} / B2B {n_b}) -> {OUT.name}  skipped_already_in_progress={skipped}"
    with open(LOG, "a") as f: f.write(line + "\n")
    print(line)

if __name__ == "__main__":
    main()
